#!/usr/bin/env python3
"""
agent_tools.py — Outils utilisables par l'agent IA BDouin.
"""

import os
import re
import io
import json
import uuid
import psycopg2
import psycopg2.extras
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = os.environ.get(
    'BDOUIN_DB',
    'postgresql://postgres:FnaPWAOtCnCLDJJbcRgkOJRvESnUHUVH@shortline.proxy.rlwy.net:33685/railway'
)
HERE = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(HERE, 'static', 'exports')
os.makedirs(EXPORTS_DIR, exist_ok=True)

MAX_ROWS_QUERY = 1000
MAX_ROWS_EXPORT = 50000


def get_conn():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)


def safe_filename(filename: str, default_ext: str = '.xlsx') -> str:
    """Sanitize filename — only alphanum, dash, underscore, dot."""
    if not filename:
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    base = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
    if not any(base.endswith(ext) for ext in ['.xlsx', '.html', '.pdf', '.csv']):
        base += default_ext
    return base[:120]


def is_safe_select(sql: str) -> bool:
    """Allow only SELECT/WITH/EXPLAIN. Block destructive."""
    cleaned = sql.strip().lower()
    forbidden = ('insert', 'update', 'delete', 'drop', 'truncate', 'alter',
                 'create', 'grant', 'revoke', 'copy ')
    if any(f in cleaned for f in forbidden):
        return False
    return cleaned.startswith('select') or cleaned.startswith('with') or cleaned.startswith('explain')


# ─── Tool: query_db ───────────────────────────────────────────────────
def query_db(sql: str, max_rows: int = 100):
    """Read-only SQL. Returns columns + rows (capped)."""
    if not is_safe_select(sql):
        return {'error': 'Only SELECT/WITH/EXPLAIN queries allowed'}
    max_rows = min(max_rows or 100, MAX_ROWS_QUERY)
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchmany(max_rows)
                cols = [d[0] for d in cur.description]
                return {
                    'columns': cols,
                    'rows': [dict(r) for r in rows],
                    'rowcount': len(rows),
                }
    except Exception as e:
        return {'error': str(e)[:400]}


# ─── Tool: generate_excel ─────────────────────────────────────────────
def generate_excel(sql: str, filename: str = None, sheet_name: str = 'Data'):
    """Run SQL → write .xlsx → return download URL.
       Returns minimal info to save tokens (no full data)."""
    if not is_safe_select(sql):
        return {'error': 'Only SELECT/WITH queries allowed'}

    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchmany(MAX_ROWS_EXPORT)
                cols = [d[0] for d in cur.description]
    except Exception as e:
        return {'error': str(e)[:400]}

    filename = safe_filename(filename, '.xlsx')
    sheet_name = re.sub(r'[^\w\s-]', '', sheet_name)[:31] or 'Data'

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    for col_idx, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=col_idx, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(cols, 1):
            v = row[col]
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False, default=str)[:32760]
            ws.cell(row=row_idx, column=col_idx, value=v)

    for col_idx, col in enumerate(cols, 1):
        col_letter = get_column_letter(col_idx)
        sample_lens = [len(str(row[col])) for row in rows[:50] if row[col] is not None]
        max_len = max(sample_lens + [len(col)], default=12)
        ws.column_dimensions[col_letter].width = min(50, max(12, max_len + 2))

    ws.freeze_panes = 'A2'

    out_path = os.path.join(EXPORTS_DIR, filename)
    wb.save(out_path)

    # Return ONLY summary, no row data (saves tokens)
    return {
        'success': True,
        'filename': filename,
        'rowcount': len(rows),
        'columns': cols,
        'download_url': f'/static/exports/{filename}',
        'size_kb': round(os.path.getsize(out_path) / 1024, 1),
    }


# ─── Tool: generate_chart ─────────────────────────────────────────────
def generate_chart(
    sql: str,
    chart_type: str = 'line',
    x_column: str = None,
    y_column: str = None,
    color_column: str = None,
    title: str = '',
    filename: str = None,
):
    """Generate Plotly HTML chart from SQL."""
    if not is_safe_select(sql):
        return {'error': 'Only SELECT/WITH queries allowed'}

    try:
        import plotly.express as px
        import pandas as pd
    except ImportError:
        return {'error': 'plotly/pandas not installed'}

    try:
        with get_conn() as conn:
            df = pd.read_sql(sql, conn)
    except Exception as e:
        return {'error': str(e)[:400]}

    if len(df) == 0:
        return {'error': 'Query returned 0 rows'}

    # Smart defaults
    if not x_column:
        x_column = df.columns[0]
    if not y_column:
        for c in df.columns:
            if c != x_column and pd.api.types.is_numeric_dtype(df[c]):
                y_column = c
                break
        if not y_column:
            y_column = df.columns[1] if len(df.columns) > 1 else df.columns[0]

    if x_column not in df.columns or y_column not in df.columns:
        return {'error': f'Column not found. Available: {list(df.columns)}'}

    try:
        kwargs = {'title': title or f'{y_column} vs {x_column}'}
        if color_column and color_column in df.columns:
            kwargs['color'] = color_column

        if chart_type == 'line':
            fig = px.line(df, x=x_column, y=y_column, markers=True, **kwargs)
        elif chart_type == 'bar':
            fig = px.bar(df, x=x_column, y=y_column, **kwargs)
        elif chart_type == 'scatter':
            fig = px.scatter(df, x=x_column, y=y_column, **kwargs)
        elif chart_type == 'area':
            fig = px.area(df, x=x_column, y=y_column, **kwargs)
        elif chart_type == 'pie':
            fig = px.pie(df, names=x_column, values=y_column, title=kwargs['title'])
        else:
            return {'error': f'Unknown chart_type: {chart_type}. Use line/bar/scatter/area/pie'}
    except Exception as e:
        return {'error': f'Chart build failed: {str(e)[:300]}'}

    fig.update_layout(
        template='plotly_dark',
        plot_bgcolor='#181b22',
        paper_bgcolor='#0f1115',
        font=dict(color='#e8eaed', family='-apple-system,sans-serif'),
        margin=dict(l=40, r=20, t=60, b=40),
    )

    filename = safe_filename(filename, '.html')
    if not filename.endswith('.html'):
        filename = filename.rsplit('.', 1)[0] + '.html'

    out_path = os.path.join(EXPORTS_DIR, filename)
    fig.write_html(out_path, include_plotlyjs='cdn', full_html=True)

    return {
        'success': True,
        'filename': filename,
        'chart_url': f'/static/exports/{filename}',
        'chart_type': chart_type,
        'rows_plotted': len(df),
        'columns_used': {'x': x_column, 'y': y_column, 'color': color_column},
        'size_kb': round(os.path.getsize(out_path) / 1024, 1),
    }


# ─── Tool: list_tables ─────────────────────────────────────────────────
def list_tables(schema: str = None):
    """List tables. If schema=None, returns gold + intelligence + clean."""
    schemas = [schema] if schema else ['gold', 'intelligence', 'clean']
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT table_schema, table_name
                    FROM information_schema.tables
                    WHERE table_schema = ANY(%s)
                    ORDER BY table_schema, table_name
                """, (schemas,))
                return {'tables': [f"{r['table_schema']}.{r['table_name']}" for r in cur.fetchall()]}
    except Exception as e:
        return {'error': str(e)}


# ─── Tool: describe_table ─────────────────────────────────────────────
def describe_table(schema: str, table: str):
    """Schema + 3 sample rows."""
    if not re.match(r'^[a-z_]+$', schema) or not re.match(r'^[a-z_0-9]+$', table):
        return {'error': 'Invalid schema/table name'}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT column_name, data_type, is_nullable
                    FROM information_schema.columns
                    WHERE table_schema=%s AND table_name=%s
                    ORDER BY ordinal_position
                """, (schema, table))
                cols = [dict(r) for r in cur.fetchall()]
                if not cols:
                    return {'error': f'{schema}.{table} not found'}

                cur.execute(f'SELECT * FROM "{schema}"."{table}" LIMIT 3')
                samples = []
                for r in cur.fetchall():
                    row = {}
                    for k, v in dict(r).items():
                        if isinstance(v, (dict, list)):
                            row[k] = json.dumps(v, ensure_ascii=False, default=str)[:200]
                        elif v is not None:
                            row[k] = str(v)[:100]
                        else:
                            row[k] = None
                    samples.append(row)

                cur.execute(f'SELECT COUNT(*) AS n FROM "{schema}"."{table}"')
                n = cur.fetchone()['n']
                return {'columns': cols, 'sample_rows': samples, 'rowcount': n}
    except Exception as e:
        return {'error': str(e)[:400]}


# ─── Tool registry ────────────────────────────────────────────────────
TOOLS_SCHEMA = [
    {
        "name": "query_db",
        "description": (
            "Execute a SELECT/WITH SQL query on the BDouin Postgres DB. "
            "Schemas available: gold.* (joined business entities), intelligence.* (features, clusters, opportunities, forecasts), "
            "clean.* (cleaned source tables), public.* (raw). "
            "Returns columns + up to 100 rows. Always use LIMIT for safety."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "sql": {"type": "string", "description": "SELECT or WITH statement"},
                "max_rows": {"type": "integer", "description": "Max rows (default 100, max 1000)"}
            },
            "required": ["sql"],
        }
    },
    {
        "name": "generate_excel",
        "description": (
            "Run SQL and produce a downloadable .xlsx file with header styling. "
            "Use when user asks for a list, export, or downloadable data. "
            "Returns file URL — does NOT include row data in response (saves tokens)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "sql": {"type": "string"},
                "filename": {"type": "string", "description": "Optional, e.g. 'top_clients.xlsx'"},
                "sheet_name": {"type": "string", "description": "Optional sheet name"},
            },
            "required": ["sql"],
        }
    },
    {
        "name": "generate_chart",
        "description": (
            "Run SQL and create an interactive Plotly HTML chart. "
            "Use for any visualization (evolution, comparison, distribution). "
            "Chart types: line (time series), bar (compare categories), scatter, area, pie."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "sql": {"type": "string"},
                "chart_type": {"type": "string", "enum": ["line", "bar", "scatter", "area", "pie"]},
                "x_column": {"type": "string"},
                "y_column": {"type": "string"},
                "color_column": {"type": "string", "description": "Optional, for grouping"},
                "title": {"type": "string"},
                "filename": {"type": "string"},
            },
            "required": ["sql", "chart_type"],
        }
    },
    {
        "name": "list_tables",
        "description": "List tables in DB schemas. Default returns gold, intelligence, clean.",
        "input_schema": {
            "type": "object",
            "properties": {
                "schema": {"type": "string", "description": "Optional: 'gold', 'intelligence', 'clean', 'public'"},
            },
        }
    },
    {
        "name": "describe_table",
        "description": "Get table columns + 3 sample rows. Use when you need to see structure of a specific table.",
        "input_schema": {
            "type": "object",
            "properties": {
                "schema": {"type": "string"},
                "table": {"type": "string"},
            },
            "required": ["schema", "table"],
        }
    },
]


TOOL_FUNCTIONS = {
    'query_db': query_db,
    'generate_excel': generate_excel,
    'generate_chart': generate_chart,
    'list_tables': list_tables,
    'describe_table': describe_table,
}


def call_tool(name: str, **kwargs):
    fn = TOOL_FUNCTIONS.get(name)
    if not fn:
        return {'error': f'Unknown tool: {name}'}
    try:
        return fn(**kwargs)
    except Exception as e:
        return {'error': f'Tool error: {str(e)[:300]}'}


# CLI test
if __name__ == '__main__':
    print('=== query_db ===')
    print(query_db("SELECT segment, COUNT(*) FROM intelligence.features_user GROUP BY 1 ORDER BY 2 DESC"))
    print('\n=== generate_excel ===')
    print(generate_excel(
        "SELECT primary_email, segment, monetary FROM intelligence.features_user WHERE segment='champion' ORDER BY monetary DESC LIMIT 50",
        filename='test_champions.xlsx'
    ))
    print('\n=== generate_chart ===')
    print(generate_chart(
        "SELECT period::text AS month, ca_net_eur FROM public.sofiadis_b2b_monthly ORDER BY period",
        chart_type='line',
        x_column='month', y_column='ca_net_eur',
        title='CA B2B Sofiadis mensuel',
        filename='test_b2b.html'
    ))
