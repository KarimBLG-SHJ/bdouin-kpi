#!/usr/bin/env python3
"""
collect_drive_content.py — Aspire le CONTENU des Google Sheets, Docs, et fichiers
                            Excel/Word/PDF stockés dans drive_files.

Pré-requis :
  - drive_files déjà peuplé (collect_drive_oauth.py)
  - /tmp/drive_token.json valide

Stockage : ajoute colonne content_text + content_data (JSONB) à drive_files.
"""

import os
import io
import json
import time
import psycopg2
from psycopg2.extras import execute_values
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload

DB_URL = 'postgresql://postgres:FnaPWAOtCnCLDJJbcRgkOJRvESnUHUVH@shortline.proxy.rlwy.net:33685/railway'
TOKEN  = '/tmp/drive_token.json'
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

# Limit content size to avoid blowing up DB
MAX_TEXT = 100_000   # 100KB text per file
MAX_FILE_SIZE = 50 * 1024 * 1024  # don't download >50MB


def get_creds():
    creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(TOKEN, 'w') as f:
            f.write(creds.to_json())
    return creds


def ensure_columns(cur):
    cur.execute("""
        ALTER TABLE drive_files
          ADD COLUMN IF NOT EXISTS content_text TEXT,
          ADD COLUMN IF NOT EXISTS content_data JSONB,
          ADD COLUMN IF NOT EXISTS content_collected_at TIMESTAMP
    """)
    cur.connection.commit()


def parse_xlsx_bytes(data):
    """Parse Excel bytes into list of {sheet, rows}."""
    import openpyxl
    import tempfile
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(data); tmp = f.name
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
        sheets = []
        text = []
        for ws in wb.worksheets:
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 5000:  # cap rows per sheet
                    break
                cleaned = [str(c) if c is not None else '' for c in row]
                rows.append(cleaned)
                text.append(' | '.join(cleaned))
            sheets.append({'name': ws.title, 'rows': rows[:5000]})
        return sheets, '\n'.join(text)[:MAX_TEXT]
    finally:
        os.unlink(tmp)


def parse_xls_bytes(data):
    """Parse old .xls bytes."""
    import xlrd
    import tempfile
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        f.write(data); tmp = f.name
    try:
        wb = xlrd.open_workbook(tmp)
        sheets = []
        text = []
        for ws in wb.sheets():
            rows = []
            for i in range(min(ws.nrows, 5000)):
                row = []
                for j in range(ws.ncols):
                    cell = ws.cell(i, j)
                    row.append(str(cell.value))
                rows.append(row)
                text.append(' | '.join(row))
            sheets.append({'name': ws.name, 'rows': rows})
        return sheets, '\n'.join(text)[:MAX_TEXT]
    finally:
        os.unlink(tmp)


def export_native(svc, file_id, mime):
    """Export Google native files (Sheets, Docs) to xlsx/text."""
    if mime == 'application/vnd.google-apps.spreadsheet':
        export_mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif mime == 'application/vnd.google-apps.document':
        export_mime = 'text/plain'
    elif mime == 'application/vnd.google-apps.presentation':
        export_mime = 'text/plain'
    else:
        return None

    req = svc.files().export_media(fileId=file_id, mimeType=export_mime)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return fh.getvalue()


def download_binary(svc, file_id):
    req = svc.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return fh.getvalue()


def main():
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor()
    ensure_columns(cur)

    creds = get_creds()
    svc = build('drive', 'v3', credentials=creds, cache_discovery=False)
    print('✓ Authenticated')

    # Files to process: parsable types, not yet collected
    cur.execute("""
        SELECT id, title, mime_type, file_extension, file_size
        FROM drive_files
        WHERE content_collected_at IS NULL
          AND mime_type IN (
            'application/vnd.google-apps.spreadsheet',
            'application/vnd.google-apps.document',
            'application/vnd.google-apps.presentation',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
            'text/csv',
            'text/plain'
          )
          AND (file_size IS NULL OR file_size < %s)
        ORDER BY modified_time DESC NULLS LAST
    """, (MAX_FILE_SIZE,))
    files = cur.fetchall()
    print(f'\n{len(files)} files to parse')

    success = 0
    failed = 0
    for i, (fid, title, mime, ext, size) in enumerate(files):
        try:
            data = None
            if mime.startswith('application/vnd.google-apps'):
                data = export_native(svc, fid, mime)
            else:
                data = download_binary(svc, fid)

            if not data:
                continue

            content_data = None
            content_text = ''

            if mime in ('application/vnd.google-apps.spreadsheet',
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
                sheets, text = parse_xlsx_bytes(data)
                content_data = sheets
                content_text = text
            elif mime == 'application/vnd.ms-excel':
                sheets, text = parse_xls_bytes(data)
                content_data = sheets
                content_text = text
            elif mime in ('application/vnd.google-apps.document',
                          'application/vnd.google-apps.presentation',
                          'text/plain', 'text/csv'):
                content_text = data.decode('utf-8', errors='replace')[:MAX_TEXT]

            cur.execute("""
                UPDATE drive_files
                SET content_text = %s,
                    content_data = %s,
                    content_collected_at = NOW()
                WHERE id = %s
            """, (
                content_text.replace('\x00', '') if content_text else None,
                json.dumps(content_data, ensure_ascii=False, default=str).replace('\x00','') if content_data else None,
                fid,
            ))
            conn.commit()
            success += 1
            if (i+1) % 20 == 0:
                print(f'  {i+1}/{len(files)} | ✓ {success} ✗ {failed}')
        except Exception as e:
            conn.rollback()
            failed += 1
            err = str(e)[:80]
            cur.execute("""
                UPDATE drive_files SET content_collected_at = NOW()
                WHERE id = %s
            """, (fid,))  # mark as processed even if failed, to avoid re-trying
            conn.commit()
            if failed <= 10 or failed % 50 == 0:
                print(f'  ✗ {title[:50]}: {err}')

        time.sleep(0.1)

    print(f'\n=== DONE ===')
    print(f'  Success: {success}')
    print(f'  Failed:  {failed}')
    cur.execute('SELECT COUNT(*) FROM drive_files WHERE content_text IS NOT NULL')
    print(f'  Total parsed in DB: {cur.fetchone()[0]}')
    conn.close()


if __name__ == '__main__':
    main()
