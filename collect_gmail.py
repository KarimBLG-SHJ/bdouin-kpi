#!/usr/bin/env python3
"""
collect_gmail.py — Collecte TOUT Gmail karim80080@gmail.com
Corps des emails + pièces jointes → Railway Postgres raw_documents

Usage:
    python3 collect_gmail.py --password "xxxx xxxx xxxx xxxx"
    python3 collect_gmail.py --password "xxxx xxxx xxxx xxxx" --days 365
    python3 collect_gmail.py --password "xxxx xxxx xxxx xxxx" --all

App Password : myaccount.google.com/apppasswords
"""

import imaplib
import email
import email.header
import argparse
import json
import os
import sys
import tempfile
import time
import psycopg2
import openpyxl
import xlrd
from email.utils import parsedate_to_datetime
from datetime import datetime, timedelta

# ── Config ────────────────────────────────────────────────────────────────────

EMAIL    = "karim80080@gmail.com"
IMAP_HOST = "imap.gmail.com"
IMAP_PORT = 993
DB_URL   = "postgresql://postgres:FnaPWAOtCnCLDJJbcRgkOJRvESnUHUVH@shortline.proxy.rlwy.net:33685/railway"

# Recherches IMAP à faire (tuples: label affiché, query IMAP)
SEARCHES = [
    # Sofiadis B2B (relevés de ventes)
    ("sofiadis_b2b",      'FROM "compta3@sofiadis.fr"'),
    ("sofiadis_b2b",      'FROM "compta2@sofiadis.fr"'),
    ("sofiadis_khalil",   'FROM "km@sofiadis.fr"'),
    # Sofiadis logistique (Sofiaco)
    ("sofiaco_logistics", 'FROM "mj@sofiaco.fr"'),
    ("sofiaco_logistics", 'FROM "sofiaco.fr"'),
    # IMAK invoices
    ("imak_invoices",     'FROM "imakofset.com.tr"'),
    # Internal (Norédine, Rachid)
    ("internal",          'FROM "noredineallam@gmail.com" SUBJECT "sofiadis"'),
    ("internal",          'FROM "rachid@nexv.co" SUBJECT "sofiadis"'),
]

ATTACHMENT_EXTS = {'.xlsx', '.xls', '.csv', '.pdf', '.doc', '.docx'}

# ── DB setup ──────────────────────────────────────────────────────────────────

def get_conn():
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    return conn

def ensure_table(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS gmail_raw (
            id SERIAL PRIMARY KEY,
            message_id TEXT UNIQUE,
            source_label TEXT,
            sender TEXT,
            recipient TEXT,
            subject TEXT,
            date_sent TIMESTAMP,
            body_text TEXT,
            body_html TEXT,
            attachments JSONB,   -- [{name, type, content_json or content_b64}]
            raw_headers JSONB,
            collected_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS gmail_raw_msgid ON gmail_raw(message_id);
        CREATE INDEX IF NOT EXISTS gmail_raw_sender ON gmail_raw(sender);
        CREATE INDEX IF NOT EXISTS gmail_raw_source ON gmail_raw(source_label);
        CREATE INDEX IF NOT EXISTS gmail_raw_date ON gmail_raw(date_sent);
    """)
    conn.commit()
    cur.close()

def already_collected(conn, message_id):
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM gmail_raw WHERE message_id=%s", (message_id,))
    r = cur.fetchone()
    cur.close()
    return r is not None

# ── Attachment parsing ─────────────────────────────────────────────────────────

def read_xlsx_to_json(data: bytes) -> list:
    """Parse xlsx bytes → list of {name, rows}"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(data)
            tmp = f.name
        wb = openpyxl.load_workbook(tmp, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cleaned = []
                for cell in row:
                    if hasattr(cell, 'strftime'):
                        cleaned.append(str(cell.date()))
                    else:
                        cleaned.append(cell)
                rows.append(cleaned)
            sheets.append({"name": ws.title, "rows": rows})
        os.unlink(tmp)
        return sheets
    except Exception as e:
        return [{"error": str(e)}]

def read_xls_to_json(data: bytes) -> list:
    """Parse old .xls bytes → list of {name, rows}"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
            f.write(data)
            tmp = f.name
        wb = xlrd.open_workbook(tmp)
        sheets = []
        for ws in wb.sheets():
            rows = []
            for i in range(ws.nrows):
                row = []
                for j in range(ws.ncols):
                    cell = ws.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            row.append(str(dt.date()))
                        except Exception:
                            row.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    else:
                        row.append(cell.value)
                rows.append(row)
            sheets.append({"name": ws.name, "rows": rows})
        os.unlink(tmp)
        return sheets
    except Exception as e:
        return [{"error": str(e)}]

def safe_json(obj):
    """Convert to JSON-safe string, removing null bytes"""
    s = json.dumps(obj, ensure_ascii=False, default=str)
    return s.replace('\x00', '')

# ── Email parsing ──────────────────────────────────────────────────────────────

def decode_header_str(h):
    if not h:
        return ""
    parts = email.header.decode_header(h)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            try:
                decoded.append(part.decode(charset or 'utf-8', errors='replace'))
            except Exception:
                decoded.append(part.decode('utf-8', errors='replace'))
        else:
            decoded.append(str(part))
    return ' '.join(decoded)

def parse_message(raw_email: bytes, source_label: str) -> dict:
    msg = email.message_from_bytes(raw_email)

    message_id = msg.get('Message-ID', '').strip()
    sender     = decode_header_str(msg.get('From', ''))
    recipient  = decode_header_str(msg.get('To', ''))
    subject    = decode_header_str(msg.get('Subject', ''))

    try:
        date_sent = parsedate_to_datetime(msg.get('Date', '')).replace(tzinfo=None)
    except Exception:
        date_sent = None

    body_text = ""
    body_html = ""
    attachments = []

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            disposition  = str(part.get('Content-Disposition', ''))
            filename_raw = part.get_filename()

            if filename_raw:
                filename = decode_header_str(filename_raw)
                ext = os.path.splitext(filename)[1].lower()
                payload = part.get_payload(decode=True)
                if payload:
                    att = {"name": filename, "type": content_type, "size": len(payload)}
                    if ext == '.xlsx':
                        att["content_json"] = read_xlsx_to_json(payload)
                    elif ext == '.xls':
                        att["content_json"] = read_xls_to_json(payload)
                    elif ext == '.csv':
                        try:
                            text = payload.decode('utf-8', errors='replace').replace('\x00', '')
                            import csv, io
                            rows = list(csv.reader(io.StringIO(text)))[:5000]
                            att["content_json"] = [{"name": "csv", "rows": rows}]
                        except Exception as e:
                            att["error"] = str(e)
                    # PDFs: store filename only (too heavy to store raw)
                    else:
                        att["note"] = "binary_not_stored"
                    attachments.append(att)
            elif content_type == 'text/plain' and 'attachment' not in disposition:
                payload = part.get_payload(decode=True)
                if payload:
                    body_text += payload.decode('utf-8', errors='replace')
            elif content_type == 'text/html' and 'attachment' not in disposition:
                payload = part.get_payload(decode=True)
                if payload:
                    body_html += payload.decode('utf-8', errors='replace')
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            if msg.get_content_type() == 'text/html':
                body_html = payload.decode('utf-8', errors='replace')
            else:
                body_text = payload.decode('utf-8', errors='replace')

    # Collect raw headers (useful ones)
    raw_headers = {
        "Message-ID": message_id,
        "From": sender,
        "To": recipient,
        "Subject": subject,
        "Date": msg.get('Date', ''),
        "CC": msg.get('CC', ''),
        "In-Reply-To": msg.get('In-Reply-To', ''),
    }

    return {
        "message_id": message_id,
        "source_label": source_label,
        "sender": sender[:500],
        "recipient": recipient[:500],
        "subject": subject[:1000],
        "date_sent": date_sent,
        "body_text": body_text[:50000],   # cap at 50k chars
        "body_html": body_html[:50000],
        "attachments": safe_json(attachments),
        "raw_headers": safe_json(raw_headers),
    }

# ── Main collector ─────────────────────────────────────────────────────────────

def collect(password: str, since_days: int = None, collect_all: bool = False):
    print(f"Connecting to {IMAP_HOST}...")
    conn_db = get_conn()
    ensure_table(conn_db)
    cur = conn_db.cursor()

    imap = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    imap.login(EMAIL, password)
    print("✓ Gmail connected")

    total_new = 0
    total_skip = 0

    for source_label, query in SEARCHES:
        print(f"\n[{source_label}] Searching: {query}")

        # Gmail IMAP: All Mail folder (French account) — needs embedded quotes
        imap.select('"[Gmail]/Tous les messages"', readonly=True)

        # Build date filter
        if collect_all:
            search_criteria = query
        elif since_days:
            since = (datetime.now() - timedelta(days=since_days)).strftime("%d-%b-%Y")
            search_criteria = f'({query} SINCE {since})'
        else:
            since = (datetime.now() - timedelta(days=365)).strftime("%d-%b-%Y")
            search_criteria = f'({query} SINCE {since})'

        try:
            status, data = imap.search(None, search_criteria)
        except Exception as e:
            print(f"  Search error: {e}")
            continue

        if status != 'OK' or not data[0]:
            print(f"  0 messages found")
            continue

        msg_ids = data[0].split()
        print(f"  Found {len(msg_ids)} messages")

        for i, msg_id in enumerate(msg_ids):
            try:
                # Fetch headers first to get Message-ID
                status, hdr_data = imap.fetch(msg_id, '(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)])')
                if status != 'OK':
                    continue

                hdr_raw = hdr_data[0][1]
                hdr_msg = email.message_from_bytes(hdr_raw)
                msgid = hdr_msg.get('Message-ID', f'no-id-{msg_id.decode()}').strip()

                if already_collected(conn_db, msgid):
                    total_skip += 1
                    continue

                # Fetch full message
                status, full_data = imap.fetch(msg_id, '(RFC822)')
                if status != 'OK' or not full_data[0]:
                    continue

                raw_email = full_data[0][1]
                parsed = parse_message(raw_email, source_label)

                cur.execute("""
                    INSERT INTO gmail_raw
                        (message_id, source_label, sender, recipient, subject, date_sent,
                         body_text, body_html, attachments, raw_headers)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (message_id) DO UPDATE SET
                        attachments = EXCLUDED.attachments
                    WHERE gmail_raw.attachments::text LIKE '%"error"%'
                       OR gmail_raw.attachments::text LIKE '%"rows": []%'
                       OR gmail_raw.attachments::text = '[]'
                """, (
                    parsed["message_id"], parsed["source_label"],
                    parsed["sender"], parsed["recipient"],
                    parsed["subject"], parsed["date_sent"],
                    parsed["body_text"], parsed["body_html"],
                    parsed["attachments"], parsed["raw_headers"],
                ))
                conn_db.commit()
                total_new += 1

                if total_new % 10 == 0:
                    print(f"  ... {i+1}/{len(msg_ids)} | {total_new} new stored")

            except Exception as e:
                conn_db.rollback()
                print(f"  ✗ msg {msg_id}: {e}")
                continue

    imap.logout()
    cur.execute("SELECT COUNT(*) FROM gmail_raw")
    total = cur.fetchone()[0]
    conn_db.close()

    print(f"\n{'='*50}")
    print(f"Done! New: {total_new} | Skipped (already stored): {total_skip}")
    print(f"Total gmail_raw: {total}")

# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--password', required=True, help='Gmail App Password (16 chars)')
    parser.add_argument('--days', type=int, default=None, help='Only emails from last N days')
    parser.add_argument('--all', action='store_true', help='Collect ALL emails (no date filter)')
    args = parser.parse_args()

    collect(
        password=args.password,
        since_days=args.days,
        collect_all=args.all,
    )
