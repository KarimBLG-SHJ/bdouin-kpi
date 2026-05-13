#!/usr/bin/env python3
"""
pipeline_23_audience.py — audience.signals + audience.persons

Construit le graphe unifié des individus BDouin cross-canal :
  - 4 fichiers concours (Telegram, Quiz Awlad, Foulane, Hoopow)
  - Subscribe + opens + clicks MailerLite
  - Commandes PrestaShop (gold.b2c_orders)
  - Comments Instagram (handle conservé, pas matché email pour v0)

Sorties :
  audience.signals  — log append-only (email OU external_handle)
  audience.persons  — vue agrégée + 5 sub-scores + composite + tier
"""

import os
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

import psycopg2
from psycopg2.extras import execute_values, Json
import openpyxl

DB_URL = os.environ.get(
    'BDOUIN_DB',
    'postgresql://postgres:FnaPWAOtCnCLDJJbcRgkOJRvESnUHUVH@shortline.proxy.rlwy.net:33685/railway',
)

# Fichiers concours stockés à la racine du repo
ROOT = Path(__file__).resolve().parent.parent

CONTEST_TELEGRAM = ROOT / 'Jeux concours telegram.csv'
CONTEST_QUIZ     = ROOT / 'participants au concours awlad quiz de oct 2025.csv'
CONTEST_FOULANE  = ROOT / 'Concours Foulane Makkah Réponses.xlsx'
CONTEST_HOOPOW   = ROOT / 'Concours Hoopow Bdouin Réponses.xlsx'


DDL = """
CREATE SCHEMA IF NOT EXISTS audience;

CREATE TABLE IF NOT EXISTS audience.signals (
    id              BIGSERIAL PRIMARY KEY,
    email           TEXT,
    external_handle TEXT,
    source          TEXT NOT NULL,
    signal_type     TEXT NOT NULL,
    occurred_at     TIMESTAMPTZ,
    value           NUMERIC,
    metadata        JSONB,
    ingested_at     TIMESTAMPTZ DEFAULT NOW(),
    CHECK (email IS NOT NULL OR external_handle IS NOT NULL)
);
CREATE INDEX IF NOT EXISTS signals_email_idx       ON audience.signals (email);
CREATE INDEX IF NOT EXISTS signals_handle_idx      ON audience.signals (external_handle);
CREATE INDEX IF NOT EXISTS signals_src_type_idx    ON audience.signals (source, signal_type);
CREATE INDEX IF NOT EXISTS signals_occurred_idx    ON audience.signals (occurred_at);

DROP TABLE IF EXISTS audience.persons;
CREATE TABLE audience.persons (
    email                 TEXT PRIMARY KEY,
    score_monetary        NUMERIC(5,2),
    score_frequency       NUMERIC(5,2),
    score_recency         NUMERIC(5,2),
    score_advocacy        NUMERIC(5,2),
    score_breadth         NUMERIC(5,2),
    score_lifetime        NUMERIC(5,2),
    score_composite       NUMERIC(6,2),
    percentile            NUMERIC(5,2),
    tier                  TEXT,
    is_dormant            BOOLEAN,
    channels_touched      INTEGER,
    last_signal_at        TIMESTAMPTZ,
    total_signals         INTEGER,
    sources               TEXT[],
    computed_at           TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS persons_tier_idx        ON audience.persons (tier);
CREATE INDEX IF NOT EXISTS persons_composite_idx   ON audience.persons (score_composite DESC);
CREATE INDEX IF NOT EXISTS persons_dormant_idx     ON audience.persons (is_dormant);
"""


def _clean_email(value):
    if value is None:
        return None
    e = str(value).strip().lower()
    if '@' not in e or len(e) < 5 or ' ' in e:
        return None
    return e


def _parse_ts(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _bulk_insert(cur, rows):
    if not rows:
        return
    execute_values(
        cur,
        """INSERT INTO audience.signals
           (email, external_handle, source, signal_type, occurred_at, value, metadata)
           VALUES %s""",
        rows,
        template="(%s,%s,%s,%s,%s,%s,%s)",
        page_size=2000,
    )


# ---------- Ingestion des concours ----------

def ingest_contest_csv(cur, source, path):
    """Telegram + Quiz Awlad — UTF-16 LE, séparateur tab.
    Colonnes : Email, Emails sent, Opened, Clicked, Subscribed, Location, ...
    """
    rows = []
    with open(path, 'r', encoding='utf-16') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for line in reader:
            email = _clean_email(line.get('Email'))
            if not email:
                continue
            ts = _parse_ts(line.get('Subscribed') or line.get('Signup Timestamp'))
            try:
                opens = int((line.get('Opened') or '0').strip() or 0)
            except ValueError:
                opens = 0
            try:
                clicks = int((line.get('Clicked') or '0').strip() or 0)
            except ValueError:
                clicks = 0
            try:
                sent = int((line.get('Emails sent') or '0').strip() or 0)
            except ValueError:
                sent = 0

            rows.append((email, None, source, 'contest_participation', ts, None,
                         Json({'sent': sent, 'opens': opens, 'clicks': clicks})))
            if opens > 0:
                rows.append((email, None, source, 'open', ts, opens,
                             Json({'contest': True})))
            if clicks > 0:
                rows.append((email, None, source, 'click', ts, clicks,
                             Json({'contest': True})))
    _bulk_insert(cur, rows)
    return len(rows)


def ingest_contest_foulane(cur, path):
    """Foulane XLSX — 6 cols : Horodateur, intérêt T11, lu T10, tirage, email, verbatim."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 5:
            continue
        ts = _parse_ts(r[0])
        interet_t11 = r[1]
        lu_t10 = r[2]
        tirage = r[3]
        email = _clean_email(r[4])
        verbatim = r[5] if len(r) > 5 else None
        if not email:
            continue
        rows.append((email, None, 'contest_foulane', 'contest_participation', ts, None,
                     Json({'interet_t11': interet_t11, 'lu_t10': lu_t10, 'tirage': tirage})))
        if verbatim and str(verbatim).strip():
            rows.append((email, None, 'contest_foulane', 'verbatim', ts, None,
                         Json({'text': str(verbatim).strip()[:1000]})))
    wb.close()
    _bulk_insert(cur, rows)
    return len(rows)


def ingest_contest_hoopow(cur, path):
    """Hoopow XLSX — 9 cols : Horodateur, muslim_show_anciennete, foulane, hoopow,
    email, verbatim, hoopow2, hoopow3, email_alt.
    Email a migré de col 4 à col 8 au cours du formulaire — on prend le premier non-null.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 5:
            continue
        ts = _parse_ts(r[0])
        muslim = r[1]
        foulane = r[2]
        hoopow_v1 = r[3]
        email = _clean_email(r[4]) or _clean_email(r[8] if len(r) > 8 else None)
        verbatim = r[5] if len(r) > 5 else None
        hoopow_v2 = r[6] if len(r) > 6 else None
        hoopow_v3 = r[7] if len(r) > 7 else None
        if not email:
            continue
        rows.append((email, None, 'contest_hoopow', 'contest_participation', ts, None,
                     Json({
                         'muslim_show_anciennete': muslim,
                         'foulane': foulane,
                         'hoopow_v1': hoopow_v1,
                         'hoopow_v2': hoopow_v2,
                         'hoopow_v3': hoopow_v3,
                     })))
        if verbatim and str(verbatim).strip():
            rows.append((email, None, 'contest_hoopow', 'verbatim', ts, None,
                         Json({'text': str(verbatim).strip()[:1000]})))
    wb.close()
    _bulk_insert(cur, rows)
    return len(rows)


# ---------- Ingestion DB ----------

def ingest_mailerlite_subscribe(cur):
    cur.execute("""
        INSERT INTO audience.signals (email, source, signal_type, occurred_at, metadata)
        SELECT
            lower(trim(email)),
            'mailerlite',
            'subscribe',
            date_subscribe,
            jsonb_build_object('status', status, 'country', country)
        FROM public.ml_subscribers
        WHERE email IS NOT NULL
          AND email LIKE '%@%'
          AND length(email) >= 5
          AND date_subscribe IS NOT NULL
    """)
    return cur.rowcount


def ingest_mailerlite_opens(cur):
    """Opens par (campaign × subscriber). Pas de timestamp d'open — on prend
    ml_campaigns.date_send comme proxy. value = open_count."""
    cur.execute("""
        INSERT INTO audience.signals (email, source, signal_type, occurred_at, value, metadata)
        SELECT
            lower(trim(o.email)),
            'mailerlite',
            'open',
            c.date_send,
            o.open_count,
            jsonb_build_object('campaign_id', o.campaign_id, 'campaign_name', c.name)
        FROM public.ml_campaign_opens o
        LEFT JOIN clean.ml_campaigns c ON c.id = o.campaign_id
        WHERE o.email IS NOT NULL
          AND o.email LIKE '%@%'
          AND length(o.email) >= 5
    """)
    return cur.rowcount


def ingest_mailerlite_clicks(cur):
    cur.execute("""
        INSERT INTO audience.signals (email, source, signal_type, occurred_at, value, metadata)
        SELECT
            lower(trim(cl.email)),
            'mailerlite',
            'click',
            c.date_send,
            cl.click_count,
            jsonb_build_object('campaign_id', cl.campaign_id, 'campaign_name', c.name)
        FROM public.ml_campaign_clicks cl
        LEFT JOIN clean.ml_campaigns c ON c.id = cl.campaign_id
        WHERE cl.email IS NOT NULL
          AND cl.email LIKE '%@%'
          AND length(cl.email) >= 5
    """)
    return cur.rowcount


def ingest_prestashop_orders(cur):
    cur.execute("""
        INSERT INTO audience.signals (email, source, signal_type, occurred_at, value, metadata)
        SELECT
            lower(trim(email)),
            'prestashop',
            'purchase',
            ordered_at,
            total_paid_eur,
            jsonb_build_object('order_id', order_id, 'reference', reference)
        FROM gold.b2c_orders
        WHERE email IS NOT NULL
          AND email LIKE '%@%'
          AND length(email) >= 5
          AND ordered_at IS NOT NULL
          AND is_valid = true
    """)
    return cur.rowcount


def ingest_instagram_comments(cur):
    """Comments IG — pas de matching email pour v0. external_handle = username."""
    cur.execute("""
        INSERT INTO audience.signals (external_handle, source, signal_type, occurred_at, metadata)
        SELECT
            username,
            'instagram',
            'comment',
            "timestamp",
            jsonb_build_object(
                'post_id', post_id,
                'text', LEFT(COALESCE(text_clean, text), 500),
                'likes', like_count
            )
        FROM clean.meta_ig_comments
        WHERE username IS NOT NULL
          AND username <> ''
    """)
    return cur.rowcount


# ---------- Calcul des scores + persons ----------

def compute_persons(cur):
    """Calcule les 6 sub-scores (M/F/R/A/B/L), le composite, le percentile, le tier
    et le flag is_dormant.

    - Frequency : 12m glissants (récent)
    - Lifetime : all-time, log-scale (intérêt historique, coef plus faible)
    - Recency : décroissance expo demi-vie 90j
    - is_dormant : transverse (>540j sans signal), n'écrase pas le tier
    - Tier : percentile global → un ex-actif garde sa place dans l'échelle
    """
    cur.execute("""
        WITH agg AS (
            SELECT
                email,
                COALESCE(SUM(CASE WHEN signal_type='purchase' THEN value ELSE 0 END), 0) AS monetary_raw,
                COUNT(*) FILTER (
                    WHERE signal_type IN ('open','click','purchase','comment','contest_participation','verbatim')
                      AND occurred_at >= NOW() - INTERVAL '12 months'
                ) AS frequency_raw,
                COUNT(*) FILTER (
                    WHERE signal_type IN ('open','click','purchase','comment','contest_participation','verbatim')
                ) AS lifetime_raw,
                EXTRACT(EPOCH FROM (
                    NOW() - MAX(occurred_at) FILTER (
                        WHERE signal_type IN ('open','click','purchase','comment','contest_participation','verbatim')
                    )
                )) / 86400.0 AS days_since,
                COUNT(*) FILTER (
                    WHERE signal_type IN ('verbatim','comment','review','mention','contest_participation')
                      AND occurred_at >= NOW() - INTERVAL '24 months'
                ) AS advocacy_raw,
                COUNT(DISTINCT CASE
                    WHEN signal_type IN ('open','click','purchase','comment','contest_participation','verbatim','review','mention')
                    THEN source
                END) AS breadth_raw,
                MAX(occurred_at) AS last_signal_at,
                COUNT(*) AS total_signals,
                array_agg(DISTINCT source) AS sources
            FROM audience.signals
            WHERE email IS NOT NULL
            GROUP BY email
        ),
        scaled AS (
            SELECT *,
                LEAST(100.0, GREATEST(0.0,
                    LN(GREATEST(monetary_raw, 0) + 1) / LN(1001) * 100
                ))::numeric(5,2) AS score_monetary,
                LEAST(100.0,
                    LN(frequency_raw + 1) / LN(201) * 100
                )::numeric(5,2) AS score_frequency,
                (CASE
                    WHEN days_since IS NULL THEN 0
                    ELSE 100.0 * EXP(-LN(2) * days_since / 90.0)
                END)::numeric(5,2) AS score_recency,
                LEAST(100.0,
                    LN(advocacy_raw + 1) / LN(51) * 100
                )::numeric(5,2) AS score_advocacy,
                LEAST(100.0, breadth_raw * 100.0 / 6.0)::numeric(5,2) AS score_breadth,
                LEAST(100.0,
                    LN(lifetime_raw + 1) / LN(501) * 100
                )::numeric(5,2) AS score_lifetime
            FROM agg
        ),
        composite AS (
            SELECT *,
                (0.25 * score_monetary
                 + 0.20 * score_frequency
                 + 0.15 * score_recency
                 + 0.15 * score_advocacy
                 + 0.10 * score_breadth
                 + 0.15 * score_lifetime)::numeric(6,2) AS score_composite
            FROM scaled
        ),
        ranked AS (
            SELECT *,
                (PERCENT_RANK() OVER (ORDER BY score_composite) * 100)::numeric(5,2) AS percentile
            FROM composite
        )
        INSERT INTO audience.persons (
            email,
            score_monetary, score_frequency, score_recency, score_advocacy, score_breadth, score_lifetime,
            score_composite, percentile, tier, is_dormant,
            channels_touched, last_signal_at, total_signals, sources
        )
        SELECT
            email,
            score_monetary, score_frequency, score_recency, score_advocacy, score_breadth, score_lifetime,
            score_composite, percentile,
            -- Tier par seuils fixes sur le composite (calibrés empiriquement v0,
            -- voir distribution dans le snapshot 2026-05-13).
            CASE
                WHEN score_composite >= 25 THEN 'diamond'
                WHEN score_composite >= 15 THEN 'platinum'
                WHEN score_composite >=  8 THEN 'gold'
                WHEN score_composite >=  3 THEN 'silver'
                ELSE 'bronze'
            END AS tier,
            (days_since IS NULL OR days_since > 540) AS is_dormant,
            breadth_raw, last_signal_at, total_signals, sources
        FROM ranked
    """)
    return cur.rowcount


# ---------- Main ----------

def main():
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor()

    print("[1/5] DDL...")
    cur.execute(DDL)

    print("[2/5] TRUNCATE audience.signals...")
    cur.execute("TRUNCATE audience.signals RESTART IDENTITY")

    print("[3/5] Ingestion concours...")
    for label, path, fn in [
        ('telegram',   CONTEST_TELEGRAM, lambda c: ingest_contest_csv(c, 'contest_telegram', CONTEST_TELEGRAM)),
        ('quiz_awlad', CONTEST_QUIZ,     lambda c: ingest_contest_csv(c, 'contest_quiz_awlad', CONTEST_QUIZ)),
        ('foulane',    CONTEST_FOULANE,  lambda c: ingest_contest_foulane(c, CONTEST_FOULANE)),
        ('hoopow',     CONTEST_HOOPOW,   lambda c: ingest_contest_hoopow(c, CONTEST_HOOPOW)),
    ]:
        if not path.exists():
            print(f"  [skip] {label} : fichier manquant ({path})")
            continue
        n = fn(cur)
        print(f"  {label:12s} : {n:>7,} signaux")

    print("[4/5] Ingestion DB...")
    print(f"  ml_subscribe    : {ingest_mailerlite_subscribe(cur):>7,}")
    print(f"  ml_opens        : {ingest_mailerlite_opens(cur):>7,}")
    print(f"  ml_clicks       : {ingest_mailerlite_clicks(cur):>7,}")
    print(f"  presta_purchase : {ingest_prestashop_orders(cur):>7,}")
    print(f"  ig_comments     : {ingest_instagram_comments(cur):>7,}")

    print("[5/5] Calcul audience.persons...")
    n = compute_persons(cur)
    print(f"  persons         : {n:>7,}")

    conn.commit()

    cur.execute("""
        SELECT tier, COUNT(*) FROM audience.persons GROUP BY tier ORDER BY COUNT(*) DESC
    """)
    print("\nDistribution tiers :")
    for tier, count in cur.fetchall():
        print(f"  {tier:10s} : {count:>7,}")

    cur.close()
    conn.close()
    print("\nDONE")


if __name__ == '__main__':
    main()
