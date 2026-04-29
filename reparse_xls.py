#!/usr/bin/env python3
"""
reparse_xls.py — Re-fetch and re-parse emails where .xls attachment failed (openpyxl error).
Fetches those message_ids from Gmail IMAP, re-parses with xlrd, updates DB.

Usage:
    python3 reparse_xls.py --password "xxxx xxxx xxxx xxxx"
"""

import imaplib
import email
import email.header
import argparse
import json
import os
import sys
import tempfile
import psycopg2
import openpyxl
import xlrd
from email.utils import parsedate_to_datetime

EMAIL     = "karim80080@gmail.com"
IMAP_HOST = "imap.gmail.com"
IMAP_PORT = 993
DB_URL    = "postgresql://postgres:FnaPWAOtCnCLDJJbcRgkOJRvESnUHUVH@shortline.proxy.rlwy.net:33685/railway"


def read_xlsx_to_json(data):
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(data); tmp = f.name
        wb = openpyxl.load_workbook(tmp, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cleaned = [str(c.date()) if hasattr(c, 'strftime') else c for c in row]
                rows.append(cleaned)
            sheets.append({"name": ws.title, "rows": rows})
        os.unlink(tmp)
        return sheets
    except Exception as e:
        return [{"error": str(e)}]


def read_xls_to_json(data):
    try:
        with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
            f.write(data); tmp = f.name
        wb = xlrd.open_workbook(tmp)
        sheets = []
        for ws in wb.sheets():
            rows = []
            for i in range(ws.nrows):
                row = []
                for j in range(ws.ncols):
                    cell = ws.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                            row.append(str(dt.date()))
                        except Exception:
                            row.append(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    else:
                        row.append(cell.value)
                rows.append(row)
            sheets.append({"name": ws.name, "rows": rows})
        os.unlink(tmp)
        return sheets
    except Exception as e:
        return [{"error": str(e)}]


def safe_json(obj):
    s = json.dumps(obj, ensure_ascii=False, default=str)
    return s.replace('\x00', '')


def decode_header_str(h):
    if not h: return ""
    parts = email.header.decode_header(h)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            try: decoded.append(part.decode(charset or 'utf-8', errors='replace'))
            except: decoded.append(part.decode('utf-8', errors='replace'))
        else:
            decoded.append(str(part))
    return ' '.join(decoded)


def parse_attachments(msg):
    attachments = []
    body_text = ""
    body_html = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            disposition  = str(part.get('Content-Disposition', ''))
            filename_raw = part.get_filename()
            if filename_raw:
                filename = decode_header_str(filename_raw)
                ext = os.path.splitext(filename)[1].lower()
                payload = part.get_payload(decode=True)
                if payload:
                    att = {"name": filename, "type": content_type, "size": len(payload)}
                    if ext == '.xlsx':
                        att["content_json"] = read_xlsx_to_json(payload)
                    elif ext == '.xls':
                        att["content_json"] = read_xls_to_json(payload)
                    elif ext == '.csv':
                        try:
                            text = payload.decode('utf-8', errors='replace').replace('\x00', '')
                            import csv, io
                            att["content_json"] = [{"name": "csv", "rows": list(csv.reader(io.StringIO(text)))[:5000]}]
                        except Exception as e:
                            att["error"] = str(e)
                    else:
                        att["note"] = "binary_not_stored"
                    attachments.append(att)
            elif content_type == 'text/plain' and 'attachment' not in disposition:
                p = part.get_payload(decode=True)
                if p: body_text += p.decode('utf-8', errors='replace')
            elif content_type == 'text/html' and 'attachment' not in disposition:
                p = part.get_payload(decode=True)
                if p: body_html += p.decode('utf-8', errors='replace')
    else:
        p = msg.get_payload(decode=True)
        if p:
            if msg.get_content_type() == 'text/html':
                body_html = p.decode('utf-8', errors='replace')
            else:
                body_text = p.decode('utf-8', errors='replace')
    return attachments, body_text, body_html


def reparse(password):
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor()

    # Get all message_ids with xls parse errors
    cur.execute("""
        SELECT message_id FROM gmail_raw
        WHERE attachments::text LIKE '%"error": "File is not a zip file"%'
        ORDER BY date_sent
    """)
    bad_msgids = {r[0] for r in cur.fetchall()}
    print(f"Emails to re-parse: {len(bad_msgids)}")

    imap = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    imap.login(EMAIL, password)
    print("✓ Gmail connected")
    imap.select('"[Gmail]/Tous les messages"', readonly=True)

    updated = 0

    for msgid in sorted(bad_msgids):
        # Search by Message-ID header directly
        msgid_clean = msgid.strip('<>')
        status, data = imap.search(None, f'HEADER Message-ID "{msgid_clean}"')
        if status != 'OK' or not data[0]:
            # Try with angle brackets
            status, data = imap.search(None, f'HEADER Message-ID "<{msgid_clean}>"')
            if status != 'OK' or not data[0]:
                print(f"  Not found: {msgid[:50]}")
                continue

        imap_ids = data[0].split()
        imap_id = imap_ids[0]  # Take first match

        status, full_data = imap.fetch(imap_id, '(RFC822)')
        if status != 'OK' or not full_data[0]:
            print(f"  Fetch failed: {msgid[:50]}")
            continue

        raw_email = full_data[0][1]
        msg = email.message_from_bytes(raw_email)
        attachments, body_text, body_html = parse_attachments(msg)

        cur.execute("""
            UPDATE gmail_raw SET attachments = %s
            WHERE message_id = %s
        """, (safe_json(attachments), msgid))
        conn.commit()
        updated += 1
        print(f"  ✓ [{updated}/{len(bad_msgids)}] {msgid[:50]}")

    imap.logout()
    print(f"\nDone! Updated: {updated}/{len(bad_msgids)} emails")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--password', required=True)
    args = parser.parse_args()
    reparse(args.password)
